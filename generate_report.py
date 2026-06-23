#!/usr/bin/env python3
"""
Generador del Informe de Seguimiento de Estimados - Xlectrical.
Lee el Excel del tracker y produce un informe HTML (versión pública enmascarada:
apellidos -> inicial, telefonos -> solo codigo de area).

Uso:
    python3 generate_report.py <ruta_excel.xlsx> <ruta_salida.html> [--logo <logo.png>]

No depende de archivos externos salvo openpyxl y (opcional) el logo PNG.
"""
import sys, base64, re, os
from datetime import date, datetime, timedelta
from collections import defaultdict, Counter
import openpyxl

# Logo en base64 (archivo logo.b64 junto a este script). Si no está, usa el membrete "XL".
try:
    LOGO_B64 = open(os.path.join(os.path.dirname(os.path.abspath(__file__)), "logo.b64")).read().strip()
except Exception:
    LOGO_B64 = ""

ACT = {"Nuevo","En seguimiento","Negociación","Esperando aprobación","Agendado","Frío"}

CSS = r"""
  :root{
    --pur:#6a0dce; --pur-d:#4a0b96; --pur-br:#8125d7; --pur-l:#f1e9fb; --pur-ll:#f9f5fe;
    --gold:#ffb000; --gold-d:#e09600; --gold-l:#fff6e3;
    --ink:#211a2e; --muted:#726b80; --line:#e9e4f2; --bg:#f6f3fb; --card:#fff;
    --red:#c0392b; --red-l:#fbecea; --amber:#b9770e; --green:#1e7a4d;
  }
  *{box-sizing:border-box}
  body{margin:0;background:var(--bg);color:var(--ink);
    font-family:'Segoe UI',-apple-system,BlinkMacSystemFont,Roboto,Helvetica,Arial,sans-serif;
    font-size:14px;line-height:1.5;-webkit-font-smoothing:antialiased}
  .wrap{max-width:1080px;margin:0 auto;padding:0 20px 60px}
  header{background:var(--card);border-bottom:3px solid var(--gold);box-shadow:0 2px 16px rgba(74,11,150,.07)}
  .hin{max-width:1080px;margin:0 auto;padding:22px 20px;display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:14px}
  .hin img{height:46px;width:auto;display:block}
  .date{text-align:right;font-size:13px;color:var(--muted)}
  .date .tag{display:inline-block;font-size:11px;font-weight:700;letter-spacing:.6px;color:var(--pur);
    background:var(--pur-l);padding:3px 10px;border-radius:20px;text-transform:uppercase}
  .date b{display:block;font-size:16px;color:var(--ink);margin-top:5px}
  h1.page{max-width:1080px;margin:26px auto 4px;padding:0 20px;font-size:22px;font-weight:800;letter-spacing:-.3px}
  .lede{max-width:1080px;margin:0 auto 24px;padding:0 20px;color:var(--muted);font-size:13.5px}
  .alert{background:var(--card);border:1px solid var(--line);border-left:5px solid var(--gold);
    border-radius:14px;padding:18px 20px;margin-bottom:22px;box-shadow:0 8px 24px rgba(74,11,150,.06);
    display:flex;gap:18px;align-items:center;flex-wrap:wrap}
  .alert .big{font-size:40px;font-weight:800;color:var(--pur);line-height:1;font-variant-numeric:tabular-nums}
  .alert .txt h2{margin:0 0 3px;font-size:16px} .alert .txt p{margin:0;color:var(--muted);font-size:13px}
  .chips{margin-left:auto;display:flex;gap:10px;flex-wrap:wrap}
  .chip{background:var(--bg);border:1px solid var(--line);border-radius:10px;padding:8px 13px;text-align:center;min-width:80px}
  .chip b{display:block;font-size:20px;font-variant-numeric:tabular-nums}
  .chip span{font-size:11px;color:var(--muted)}
  .chip.r b{color:var(--red)} .chip.a b{color:var(--gold-d)}
  .kpis{display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin-bottom:26px}
  .kpi{background:var(--card);border:1px solid var(--line);border-radius:14px;padding:16px 18px;box-shadow:0 6px 18px rgba(74,11,150,.05)}
  .kpi .lab{font-size:11.5px;color:var(--muted);text-transform:uppercase;letter-spacing:.5px;font-weight:600}
  .kpi .val{font-size:27px;font-weight:800;margin-top:6px;letter-spacing:-.5px;font-variant-numeric:tabular-nums}
  .kpi .note{font-size:12px;color:var(--muted);margin-top:3px}
  .kpi.accent{background:linear-gradient(135deg,var(--pur),var(--pur-d));border:none;color:#fff}
  .kpi.accent .lab{color:#e7d6fb} .kpi.accent .note{color:#d8c2f5}
  .kpi.gold{background:linear-gradient(135deg,#fff,var(--gold-l));border-color:#f2dba0}
  .kpi.gold .val{color:var(--gold-d)}
  section{background:var(--card);border:1px solid var(--line);border-radius:16px;padding:22px 24px;margin-bottom:22px;box-shadow:0 6px 18px rgba(74,11,150,.05)}
  .sh{display:flex;align-items:center;gap:11px;margin:0 0 4px}
  .sh h2{margin:0;font-size:16.5px;font-weight:700}
  .eyebrow{font-size:11px;font-weight:700;color:#fff;background:var(--pur);padding:3px 9px;border-radius:20px;letter-spacing:.5px}
  .eyebrow.g{background:var(--gold);color:#5a3d00}
  .sdesc{color:var(--muted);font-size:13px;margin:0 0 16px}
  table{width:100%;border-collapse:collapse;font-size:13px}
  th{text-align:left;font-size:11px;text-transform:uppercase;letter-spacing:.5px;color:var(--muted);font-weight:700;padding:0 10px 9px;border-bottom:2px solid var(--line)}
  th.num,td.num{text-align:right;font-variant-numeric:tabular-nums}
  td{padding:10px;border-bottom:1px solid var(--line);vertical-align:middle}
  tr:last-child td{border-bottom:none} tbody tr:hover{background:var(--pur-ll)}
  .cli{font-weight:600} .cli .sub{display:block;font-weight:400;font-size:11.5px;color:var(--muted)}
  .tel{font-variant-numeric:tabular-nums;color:var(--pur-d);font-weight:600;white-space:nowrap}
  .d{font-variant-numeric:tabular-nums;font-weight:700;padding:2px 8px;border-radius:7px;background:var(--bg);font-size:12px}
  .d-mid{background:var(--gold-l);color:var(--gold-d)} .d-hi{background:var(--red-l);color:var(--red)}
  .barwrap{position:relative;background:var(--bg);border-radius:7px;height:20px;min-width:120px;overflow:hidden}
  .bar{position:absolute;left:0;top:0;bottom:0;background:var(--pur);border-radius:7px;opacity:.9}
  .bar2{background:var(--green)} .bar3{background:var(--red)} .bar-gold{background:var(--gold)}
  .ag-hi{background:var(--red)} .ag-mid{background:var(--gold-d)}
  .barwrap span{position:absolute;right:8px;top:50%;transform:translateY(-50%);font-size:11px;font-weight:700;color:var(--ink)}
  .two{display:grid;grid-template-columns:1fr 1fr;gap:22px} .two section{margin-bottom:0}
  .divider{display:flex;align-items:center;gap:14px;margin:34px 0 20px}
  .divider::before,.divider::after{content:"";height:2px;flex:1;background:linear-gradient(90deg,transparent,var(--pur-l),transparent)}
  .divider span{font-size:12px;font-weight:800;letter-spacing:1.5px;color:var(--pur);text-transform:uppercase;
    background:var(--gold);color:#5a3d00;padding:5px 16px;border-radius:30px}
  .flag{background:var(--gold-l);border:1px solid #f2dba0;border-radius:12px;padding:14px 16px;margin-bottom:22px;font-size:13px;color:#7a5408}
  .flag b{color:#5e3f04}
  footer{text-align:center;color:var(--muted);font-size:12px;margin-top:30px;line-height:1.7}

  .takeaway{margin:14px 0 0;padding:12px 15px;background:var(--gold-l);border-left:3px solid var(--gold);
    border-radius:8px;font-size:13px;color:#5e3f04}
  .takeaway b{color:#4a3100}
  .cmp{display:flex;flex-direction:column;gap:16px;margin:6px 0}
  .cmp-row{display:grid;grid-template-columns:170px 1fr 120px;align-items:center;gap:14px}
  .cmp-lab{font-size:13px;font-weight:600} .cmp-lab b{color:var(--muted);font-weight:400;margin:0 4px}
  .cmp-track{background:var(--bg);border-radius:9px;height:34px;overflow:hidden}
  .cmp-fill{height:100%;display:flex;align-items:center;justify-content:flex-end;padding:0 12px;
    color:#fff;font-weight:800;font-size:14px;border-radius:9px;font-variant-numeric:tabular-nums;min-width:64px}
  .cmp-g{background:linear-gradient(90deg,#2a9d63,#1e7a4d)} .cmp-r{background:linear-gradient(90deg,#d4503f,#c0392b)}
  .cmp-med{font-size:12px;color:var(--muted);text-align:right}
  .dot{display:inline-block;width:9px;height:9px;border-radius:50%;margin-right:6px;vertical-align:middle}
  .dot-g{background:var(--green)} .dot-r{background:var(--red)}
  .two-stat{display:grid;grid-template-columns:1fr 1fr;gap:16px}
  .stat-card{border-radius:14px;padding:18px 20px;border:1px solid var(--line)}
  .stat-g{background:linear-gradient(160deg,#eef9f2,#fff);border-color:#cbe8d6}
  .stat-r{background:linear-gradient(160deg,#fdeeec,#fff);border-color:#f3cfc9}
  .stat-n{font-size:38px;font-weight:800;line-height:1;font-variant-numeric:tabular-nums}
  .stat-g .stat-n{color:var(--green)} .stat-r .stat-n{color:var(--red)}
  .stat-l{font-size:12.5px;color:var(--muted);margin-top:4px;line-height:1.35}
  .minis{display:flex;align-items:flex-end;gap:6px;height:48px;margin-top:14px}
  .mini{flex:1;display:flex;flex-direction:column;align-items:center;gap:4px;height:100%;justify-content:flex-end}
  .mini-bar{width:100%;border-radius:4px 4px 0 0;min-height:3px}
  .mini-g{background:var(--green);opacity:.8} .mini-r{background:var(--red);opacity:.8}
  .mini span{font-size:10px;color:var(--muted)}
  .rk-hi{background:var(--red)} .rk-mid{background:var(--gold-d)}
  @media(max-width:760px){.kpis{grid-template-columns:repeat(2,1fr)}.two{grid-template-columns:1fr}.chips{margin-left:0;width:100%}table{font-size:12px}.cmp-row{grid-template-columns:1fr;gap:5px}.cmp-med{text-align:left}.two-stat{grid-template-columns:1fr}}
"""

def fmt(n): return "$"+format(round(n),",")
def dt(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    return None
def mask_name(name):
    name=(name or "").strip()
    p=name.split()
    return p[0]+" "+p[1][0]+"." if len(p)>=2 else name
def mask_tel(t):
    s=str(int(t)) if isinstance(t,(int,float)) else str(t or "")
    return f"({s[:3]}) •••-••••" if len(s)==10 else "—"

def load(path):
    wb=openpyxl.load_workbook(path, data_only=True)
    sg=wb["Seguimiento"]; rows=[]
    for r in range(2, sg.max_row+1):
        cli=sg.cell(row=r,column=3).value
        if not (cli and str(cli).strip()): continue
        rows.append(dict(
            cliente=str(cli).strip(), tel=sg.cell(row=r,column=4).value,
            ciudad=sg.cell(row=r,column=6).value, servicio=sg.cell(row=r,column=7).value,
            tecnico=sg.cell(row=r,column=8).value, monto=sg.cell(row=r,column=10).value or 0,
            fecha_est=dt(sg.cell(row=r,column=11).value), ult=dt(sg.cell(row=r,column=12).value),
            prox=dt(sg.cell(row=r,column=13).value),
            estatus=(str(sg.cell(row=r,column=15).value).strip() if sg.cell(row=r,column=15).value else ""),
            prob=sg.cell(row=r,column=16).value or 0, fuente=sg.cell(row=r,column=17).value,
            razon=sg.cell(row=r,column=18).value, intentos=sg.cell(row=r,column=21).value))
    lsa=[]; lw=wb["Leads sin agendar"]
    for r in range(2, lw.max_row+1):
        c=lw.cell(row=r,column=2).value
        if not (c and str(c).strip()): continue
        fecha=dt(lw.cell(row=r,column=1).value); serv=lw.cell(row=r,column=4).value
        fuente=lw.cell(row=r,column=5).value; motivo=lw.cell(row=r,column=6).value
        prox=dt(lw.cell(row=r,column=7).value); est=lw.cell(row=r,column=8).value
        csr=lw.cell(row=r,column=9).value
        if not any([fecha,serv,fuente,motivo,est]): continue
        lsa.append(dict(cliente=str(c).strip(), tel=lw.cell(row=r,column=3).value, serv=serv,
            fuente=fuente, motivo=motivo, prox=prox, est=est, csr=csr))
    return rows, lsa

def bar(pct,cls=""):
    return f'<div class="barwrap"><div class="bar {cls}" style="width:{max(0,min(100,round(pct)))}%"></div><span>{round(pct)}%</span></div>'

def kp(lab,val,note,cls=""):
    return f'<div class="kpi {cls}"><div class="lab">{lab}</div><div class="val">{val}</div><div class="note">{note}</div></div>'

def analyze(rows, today):
    """Genera alert + KPIs + secciones de análisis para un subconjunto de estimados."""
    P=[]
    def dias(r): return (today-r["ult"]).days if r["ult"] else None
    n=len(rows)
    activos=[r for r in rows if r["estatus"] in ACT]
    ganados=[r for r in rows if r["estatus"]=="Ganado"]
    perdidos=[r for r in rows if r["estatus"]=="Perdido"]
    sin_est=[r for r in rows if r["estatus"]==""]
    pip=sum(r["monto"] for r in activos)
    pond=sum(r["monto"]*(r["prob"] or 0) for r in activos)
    pos=[r["monto"] for r in rows if isinstance(r["monto"],(int,float)) and r["monto"]>0]
    ticket=sum(pos)/len(pos) if pos else 0
    tasa=len(ganados)/n if n else 0
    venc=[r for r in activos if r["prox"] and r["prox"]<=today]
    tibios=[r for r in activos if (dias(r) or 0)>=14]
    call=[r for r in activos if (r["prox"] and r["prox"]<=today) or (dias(r) is not None and dias(r)>=3)]
    call.sort(key=lambda r:(r["monto"]*(r["prob"] or 0)), reverse=True)

    meses={1:"enero",2:"febrero",3:"marzo",4:"abril",5:"mayo",6:"junio",7:"julio",8:"agosto",9:"septiembre",10:"octubre",11:"noviembre",12:"diciembre"}
    fecha=f"{today.day} de {meses[today.month]} de {today.year}"

    def bar(pct,cls=""):
        return f'<div class="barwrap"><div class="bar {cls}" style="width:{max(0,min(100,round(pct)))}%"></div><span>{round(pct)}%</span></div>'

    P=[]
    # Llamadas de hoy (ENMASCARADO)
    cr=""
    for r in call[:15]:
        d=dias(r) or 0; dc="d-hi" if d>=14 else ("d-mid" if d>=7 else "")
        cr+=f'<tr><td class="cli">{mask_name(r["cliente"])}<span class="sub">{r["ciudad"] or ""}</span></td><td>{r["servicio"] or "—"}</td><td class="num">{fmt(r["monto"])}</td><td class="num">{round((r["prob"] or 0)*100)}%</td><td><span class="d {dc}">{d}d</span></td><td>{r["estatus"]}</td><td class="tel">{mask_tel(r["tel"])}</td></tr>'
    P.append(f'<section><div class="sh"><span class="eyebrow g">PRIORIDAD</span><h2>Llamadas de hoy</h2></div><p class="sdesc">Top 15 oportunidades activas por valor ponderado (monto × probabilidad).</p><table><thead><tr><th>Cliente</th><th>Servicio</th><th class="num">Monto</th><th class="num">Prob.</th><th>Sin contacto</th><th>Estatus</th><th>Teléfono</th></tr></thead><tbody>{cr}</tbody></table></section>')

    # Antigüedad $ en riesgo
    risk=defaultdict(lambda:[0,0])
    for r in activos:
        d=dias(r) or 0
        k="30+ días" if d>=30 else "14-29 días" if d>=14 else "7-13 días" if d>=7 else "0-6 días"
        risk[k][0]+=r["monto"]; risk[k][1]+=1
    mx=max([v[0] for v in risk.values()]+[1]); rr=""
    for k in ["0-6 días","7-13 días","14-29 días","30+ días"]:
        if k not in risk: continue
        amt,c=risk[k]; cls="ag-hi" if k=="30+ días" else ("ag-mid" if k=="14-29 días" else "")
        rr+=f'<tr><td class="cli">{k}</td><td class="num">{c}</td><td class="num">{fmt(amt)}</td><td>{bar(amt/mx*100,cls)}</td></tr>'
    P.append(f'<section><div class="sh"><span class="eyebrow">RIESGO $</span><h2>Pipeline en riesgo por antigüedad</h2></div><p class="sdesc">Cuánto dinero lleva sin contacto. Lo de 30+ días está por enfriarse.</p><table><thead><tr><th>Días sin contacto</th><th class="num"># est.</th><th class="num">$ en riesgo</th><th>Monto</th></tr></thead><tbody>{rr}</tbody></table></section>')

    # Técnico (total + puro)
    norm={"James Walkers":"James Walker"}; tec=defaultdict(lambda:{"g":0,"p":0,"d":0,"act":0,"mg":0})
    for r in rows:
        t=norm.get(r["tecnico"], r["tecnico"] or "Sin asignar")
        if t not in ("James Walker","Romario Taffe","Jesus Cardenas","David Sierra"): t="Otros / sin asignar"
        s=r["estatus"]
        if s=="Ganado": tec[t]["g"]+=1; tec[t]["mg"]+=r["monto"]
        elif s=="Perdido": tec[t]["p"]+=1
        elif s=="Descartado": tec[t]["d"]+=1
        elif s in ACT: tec[t]["act"]+=1
    tr=""
    for t in ["Jesus Cardenas","James Walker","Romario Taffe","David Sierra","Otros / sin asignar"]:
        v=tec[t]; tot=v["g"]+v["p"]+v["d"]+v["act"]; cerr=v["g"]+v["p"]+v["d"]
        tt=v["g"]/tot*100 if tot else 0; tp=v["g"]/cerr*100 if cerr else 0
        tr+=f'<tr><td class="cli">{t}</td><td class="num">{v["g"]}</td><td class="num">{v["p"]}</td><td class="num">{v["act"]}</td><td class="num">{fmt(v["mg"])}</td><td>{bar(tt)}</td><td>{bar(tp,"bar-gold")}</td></tr>'
    P.append(f'<section><div class="sh"><h2>Cierre por técnico</h2></div><p class="sdesc">Cierre total (÷ todos sus estimados) vs cierre puro (÷ solo los cerrados).</p><table><thead><tr><th>Técnico</th><th class="num">G</th><th class="num">P</th><th class="num">Act.</th><th class="num">$ Ganado</th><th>Cierre total</th><th>Cierre puro</th></tr></thead><tbody>{tr}</tbody></table></section>')

    # Fuente
    fue=defaultdict(lambda:{"total":0,"g":0})
    for r in rows:
        f=r["fuente"] or "—"; fue[f]["total"]+=1
        if r["estatus"]=="Ganado": fue[f]["g"]+=1
    fr=""
    for f in sorted(fue,key=lambda x:fue[x]["total"],reverse=True):
        v=fue[f]; cr2=v["g"]/v["total"]*100 if v["total"] else 0; nm=f if f!="—" else "Sin fuente"
        cls="bar2" if cr2>=60 else ("bar-gold" if cr2>=45 else "bar3")
        fr+=f'<tr><td class="cli">{nm}</td><td class="num">{v["total"]}</td><td class="num">{v["g"]}</td><td>{bar(cr2,cls)}</td></tr>'
    P.append(f'<section><div class="sh"><h2>Rendimiento por fuente</h2></div><p class="sdesc">De dónde vienen los leads y cuáles cierran mejor.</p><table><thead><tr><th>Fuente</th><th class="num">Total</th><th class="num">Ganados</th><th>Cierre</th></tr></thead><tbody>{fr}</tbody></table></section>')

    # Servicio
    serv=defaultdict(lambda:{"g":0,"p":0,"tot":0,"mg":0})
    for r in rows:
        s=r["servicio"] or "Sin especificar"; serv[s]["tot"]+=1
        if r["estatus"]=="Ganado": serv[s]["g"]+=1; serv[s]["mg"]+=r["monto"]
        elif r["estatus"]=="Perdido": serv[s]["p"]+=1
    srv=""
    for s in sorted(serv,key=lambda x:serv[x]["tot"],reverse=True):
        v=serv[s]
        if v["tot"]<2: continue
        cerr=v["g"]+v["p"]; cr2=v["g"]/cerr*100 if cerr else 0; avg=v["mg"]/v["g"] if v["g"] else 0
        cls="bar2" if cr2>=70 else ("bar-gold" if cr2>=45 else "bar3")
        srv+=f'<tr><td class="cli">{s}</td><td class="num">{v["tot"]}</td><td class="num">{fmt(avg)}</td><td>{bar(cr2,cls)}</td></tr>'
    P.append(f'<section><div class="sh"><h2>Rendimiento por tipo de servicio</h2></div><p class="sdesc">Qué trabajos cierran mejor y cuánto valen.</p><table><thead><tr><th>Servicio</th><th class="num">Total</th><th class="num">Ticket gan.</th><th>Cierre</th></tr></thead><tbody>{srv}</tbody></table></section>')

    # Velocidad
    spd=defaultdict(list)
    for r in rows:
        if r["estatus"] not in ("Ganado","Perdido"): continue
        if r["fecha_est"] and r["ult"]:
            sd=(r["ult"]-r["fecha_est"]).days
            k="Mismo día (0-1)" if sd<=1 else "2-3 días" if sd<=3 else "4-7 días" if sd<=7 else "8+ días"
            spd[k].append(r["estatus"]=="Ganado")
    vr=""
    for k in ["Mismo día (0-1)","2-3 días","4-7 días","8+ días"]:
        if k not in spd: continue
        v=spd[k]; cr2=sum(v)/len(v)*100
        cls="bar2" if cr2>=70 else ("bar-gold" if cr2>=50 else "bar3")
        vr+=f'<tr><td class="cli">{k}</td><td class="num">{len(v)}</td><td>{bar(cr2,cls)}</td></tr>'
    P.append(f'<section><div class="sh"><h2>Velocidad de respuesta vs cierre</h2></div><p class="sdesc">Entre más rápido se contacta tras el estimado, más alto el cierre.</p><table><thead><tr><th>Tiempo hasta 1er contacto</th><th class="num"># cerrados</th><th>Cierre</th></tr></thead><tbody>{vr}</tbody></table></section>')

    # Razones de pérdida
    raz=Counter(r["razon"] for r in perdidos if r["razon"]); tot=sum(raz.values()) or 1
    zr=""
    for k in sorted(raz,key=raz.get,reverse=True):
        zr+=f'<tr><td class="cli">{k}</td><td class="num">{raz[k]}</td><td>{bar(raz[k]/tot*100,"bar3")}</td></tr>'
    P.append(f'<section><div class="sh"><h2>Razones de pérdida</h2></div><p class="sdesc">Por qué se pierden los estimados.</p><table><thead><tr><th>Razón</th><th class="num"># perdidos</th><th>% del total</th></tr></thead><tbody>{zr}</tbody></table></section>')

    # Insights
    g=[r["monto"] for r in ganados if r["monto"]>0]; pl=[r["monto"] for r in perdidos if r["monto"]>0]
    gw=sum(g)/len(g) if g else 0; pw=sum(pl)/len(pl) if pl else 0
    att={"Ganado":[],"Perdido":[]}
    for r in rows:
        if r["estatus"] in att and isinstance(r["intentos"],(int,float)): att[r["estatus"]].append(r["intentos"])
    ig=sum(att["Ganado"])/len(att["Ganado"]) if att["Ganado"] else 0
    ip=sum(att["Perdido"])/len(att["Perdido"]) if att["Perdido"] else 0
    r30=risk.get("30+ días",[0,0])[0]
    P.append(f'<section><div class="sh"><span class="eyebrow g">INSIGHT</span><h2>Ticket promedio: ganado vs perdido</h2></div><p class="sdesc">Los trabajos que se ganan son mucho más chicos que los que se pierden.</p><div class="cmp"><div class="cmp-row"><div class="cmp-lab"><span class="dot dot-g"></span>Ganado <b>·</b> {len(g)}</div><div class="cmp-track"><div class="cmp-fill cmp-g" style="width:{round(gw/max(gw,pw,1)*100)}%">{fmt(gw)}</div></div><div class="cmp-med"></div></div><div class="cmp-row"><div class="cmp-lab"><span class="dot dot-r"></span>Perdido <b>·</b> {len(pl)}</div><div class="cmp-track"><div class="cmp-fill cmp-r" style="width:{round(pw/max(gw,pw,1)*100)}%">{fmt(pw)}</div></div><div class="cmp-med"></div></div></div><p class="takeaway">El estimado perdido promedio vale <b>{round(pw/gw,1) if gw else 0}× más</b> que el ganado. Reforzar el cierre en trabajos de alto valor.</p></section>')
    P.append(f'<section><div class="sh"><span class="eyebrow g">INSIGHT</span><h2>Persistencia: intentos antes de cerrar</h2></div><p class="sdesc">Los estimados que se ganan recibieron muchos más intentos que los que se perdieron.</p><div class="two-stat"><div class="stat-card stat-g"><div class="stat-n">{ig:.1f}</div><div class="stat-l">intentos promedio<br><b>en los GANADOS</b></div></div><div class="stat-card stat-r"><div class="stat-n">{ip:.1f}</div><div class="stat-l">intentos promedio<br><b>en los PERDIDOS</b></div></div></div><p class="takeaway">Se hacen <b>{ig/ip:.1f}× más intentos</b> en los que se ganan. El manual pide mínimo 3 antes de dar por perdido.</p></section>')
    alert=f'''<div class="alert"><div class="big">{len(call)}</div><div class="txt"><h2>estimados necesitan llamada hoy</h2><p>Activos con seguimiento vencido o sin contacto hace 3+ días.</p></div><div class="chips"><div class="chip r"><b>{len(venc)}</b><span>contacto vencido</span></div><div class="chip a"><b>{len(tibios)}</b><span>tibios (14+ días)</span></div><div class="chip"><b>{len(activos)}</b><span>activos totales</span></div></div></div>'''
    kpis=f'''<div class="kpis">{kp("Pipeline activo",fmt(pip),f"{len(activos)} estimados abiertos","accent")}{kp("Pipeline ponderado",fmt(pond),"× probabilidad")}{kp("Tasa de cierre",str(round(tasa*100))+"%",f"{len(ganados)} ganados / {n} totales","gold")}{kp("Ticket promedio",fmt(ticket),"por estimado")}</div>'''
    flag=f'''<div class="flag"><b>⚠ Calidad de datos:</b> {len(sin_est)} estimados con cliente pero sin estatus (no entran en el pipeline) · nombres de técnico duplicados por typo.</div>'''
    return alert + kpis + ''.join(P) + flag

def build(path_xlsx, today=None):
    today = today or date.today()
    rows, lsa = load(path_xlsx)
    n=len(rows)
    meses={1:"enero",2:"febrero",3:"marzo",4:"abril",5:"mayo",6:"junio",7:"julio",8:"agosto",9:"septiembre",10:"octubre",11:"noviembre",12:"diciembre"}
    fecha=f"{today.day} de {meses[today.month]} de {today.year}"
    # --- Dos alcances: todos los estimados y los creados en los últimos 30 días ---
    desde30 = today - timedelta(days=30)
    rows30 = [r for r in rows if r['fecha_est'] and desde30 <= r['fecha_est'] <= today]
    pane_all = analyze(rows, today)
    pane_30  = analyze(rows30, today)
    note_all = f'<p class="scope-note">Análisis sobre los <b>{n}</b> estimados del tracker (histórico completo).</p>'
    note_30  = f'<p class="scope-note">Análisis sobre los <b>{len(rows30)}</b> estimados creados en los últimos 30 días (desde el {desde30.day}/{desde30.month}).</p>'
    tabs = (
        '<div class="tabs">'
        '<button class="tab-btn active" data-t="all">Todos los estimados</button>'
        '<button class="tab-btn" data-t="r30">Últimos 30 días</button>'
        '</div>'
        f'<div class="tab-pane" id="pane-all">{note_all}{pane_all}</div>'
        f'<div class="tab-pane hidden" id="pane-r30">{note_30}{pane_30}</div>'
    )
    tab_js = ("<script>document.querySelectorAll('.tab-btn').forEach(function(b){"
        "b.addEventListener('click',function(){"
        "document.querySelectorAll('.tab-btn').forEach(function(x){x.classList.remove('active');});"
        "b.classList.add('active');"
        "document.getElementById('pane-all').classList.toggle('hidden', b.dataset.t!=='all');"
        "document.getElementById('pane-r30').classList.toggle('hidden', b.dataset.t!=='r30');"
        "});});</script>")

    # Leads sin agendar
    by_est=Counter(l["est"] for l in lsa if l["est"])
    pend=[l for l in lsa if l["est"] and "re-contacto" in str(l["est"]).lower()]
    vencp=[l for l in pend if l["prox"] and l["prox"]<=today]
    by_mot=Counter(l["motivo"] for l in lsa if l["motivo"])
    resc=round(by_est.get("Agendado",0)/len(lsa)*100) if lsa else 0
    def kp(lab,val,note,cls=""):
        return f'<div class="kpi {cls}"><div class="lab">{lab}</div><div class="val">{val}</div><div class="note">{note}</div></div>'
    lead_kpis=f'<div class="kpis">{kp("Leads registrados",len(lsa),"aún no son estimado","accent")}{kp("Por re-contactar",len(pend),f"{len(vencp)} vencidos","gold")}{kp("Agendados",by_est.get("Agendado",0),"pasaron a visita")}{kp("Tasa de rescate",str(resc)+"%","leads → agendado")}</div>'
    pr=""
    for l in sorted(pend,key=lambda x:(x["prox"] or date(2100,1,1))):
        venc_b='<span class="d d-hi">vencido</span>' if (l["prox"] and l["prox"]<=today) else ""
        pf=l["prox"].strftime("%d/%m") if l["prox"] else "—"
        pr+=f'<tr><td class="cli">{mask_name(l["cliente"])}<span class="sub">{l["fuente"] or ""}</span></td><td>{l["serv"] or "—"}</td><td>{l["motivo"] or "—"}</td><td>{pf} {venc_b}</td><td class="tel">{mask_tel(l["tel"])}</td></tr>'
    mxm=max(by_mot.values()) if by_mot else 1; mr=""
    for k in sorted(by_mot,key=by_mot.get,reverse=True):
        mr+=f'<tr><td class="cli">{k}</td><td class="num">{by_mot[k]}</td><td>{bar(by_mot[k]/mxm*100,"bar-gold")}</td></tr>'
    leads=f'<div class="divider"><span>Embudo previo · Leads sin agendar</span></div>{lead_kpis}<section><div class="sh"><span class="eyebrow g">ACCIÓN</span><h2>Leads pendientes de re-contacto</h2></div><p class="sdesc">Embudo ANTES del estimado. Todos con seguimiento vencido.</p><table><thead><tr><th>Cliente</th><th>Servicio</th><th>Motivo</th><th>Seguimiento</th><th>Teléfono</th></tr></thead><tbody>{pr}</tbody></table></section><section><div class="sh"><h2>Por qué no se agendaron</h2></div><p class="sdesc">Motivo principal de cada lead que no avanzó.</p><table><thead><tr><th>Motivo</th><th class="num"># leads</th><th>Distribución</th></tr></thead><tbody>{mr}</tbody></table></section>'

    logo_html = (f'<div class="logo-box"><img src="data:image/png;base64,{LOGO_B64}" alt="Xlectrical"></div>'
                 if LOGO_B64 else '<div class="logo">XL</div>')
    header=f'''<header><div class="hin"><div class="brand">{logo_html}<div><h1>Seguimiento de Estimados</h1><p>Xlectrical LLC · Pipeline de ventas</p></div></div><div class="date">Informe diario<b>{fecha}</b></div></div></header>'''
    footer=f'''<footer>Vista para socios · Datos de contacto enmascarados por privacidad<br>Generado automáticamente · {fecha} · {n} estimados · Xlectrical LLC</footer>'''
    EXTRA_CSS=".logo-box{background:#fff;border-radius:10px;padding:7px 12px;display:inline-block;line-height:0;box-shadow:0 4px 14px rgba(0,0,0,.18)}.logo-box img{height:34px;width:auto;display:block}"
    EXTRA_CSS += ".tabs{display:flex;gap:8px;margin:22px 0 16px;flex-wrap:wrap}.tab-btn{cursor:pointer;border:1px solid var(--line);background:var(--card);color:var(--muted);font-weight:700;font-size:13.5px;padding:10px 18px;border-radius:10px;font-family:inherit}.tab-btn.active{background:var(--pur);color:#fff;border-color:var(--pur)}.tab-btn:hover{border-color:var(--pur-br)}.hidden{display:none}.scope-note{font-size:12.5px;color:var(--muted);margin:0 0 12px}"
    body=header+'<div class="wrap">'+tabs+tab_js+leads+footer+'</div>'
    html=f'''<!DOCTYPE html><html lang="es"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1"><title>Seguimiento de Estimados — Xlectrical</title><style>{CSS}{EXTRA_CSS}</style></head><body>{body}</body></html>'''
    return html

if __name__=="__main__":
    xlsx=sys.argv[1]; out=sys.argv[2]
    html=build(xlsx)
    open(out,"w",encoding="utf-8").write(html)
    print(f"OK {len(html)} bytes -> {out}")
